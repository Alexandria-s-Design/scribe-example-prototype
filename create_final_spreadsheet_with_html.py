import json
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
import os

def get_github_username():
    """Get GitHub username"""
    import subprocess
    result = subprocess.run(['gh', 'api', 'user', '--jq', '.login'], capture_output=True, text=True, check=True)
    return result.stdout.strip()

def get_github_image_url(username, week_num):
    """Generate GitHub raw URL for an image"""
    return f"https://raw.githubusercontent.com/{username}/modelit-blog-images/main/images/week-{week_num:02d}-hero.jpg"

def json_to_simplified_html(week_data, image_url):
    """Convert JSON blog data to simplified HTML with embedded image"""

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{week_data['title']}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            line-height: 1.6;
            max-width: 800px;
            margin: 0 auto;
            padding: 20px;
            color: #333;
        }}
        .hero-image {{
            width: 100%;
            max-width: 100%;
            height: auto;
            margin-bottom: 30px;
        }}
        .excerpt {{
            font-size: 1.2em;
            font-style: italic;
            color: #555;
            padding: 15px;
            background-color: #f5f5f5;
            border-left: 4px solid #0066cc;
            margin: 20px 0;
        }}
        h2 {{
            color: #0066cc;
            margin-top: 30px;
        }}
        .cta-buttons {{
            text-align: center;
            margin: 40px 0;
        }}
        .cta-button {{
            display: inline-block;
            padding: 15px 30px;
            margin: 10px;
            color: white;
            text-decoration: none;
            font-weight: bold;
            border-radius: 5px;
        }}
        .cta-primary {{
            background-color: #0066cc;
        }}
        .cta-secondary {{
            background-color: #28a745;
        }}
        .footer {{
            margin-top: 40px;
            padding: 20px;
            background-color: #f0f0f0;
            border-top: 2px solid #ccc;
        }}
    </style>
</head>
<body>
    <img src="{image_url}" alt="{week_data['title']}" class="hero-image">

    <h1>{week_data['title']}</h1>

    <div class="excerpt">
        <p>{week_data['excerpt']}</p>
    </div>

    <p>{week_data['content']['opening_hook']}</p>

    <h2>{week_data['content']['section_1_title']}</h2>
    <p>{week_data['content']['section_1_text']}</p>

    <h2>{week_data['content']['section_2_title']}</h2>
    <p>{week_data['content']['section_2_text']}</p>

    <h2>{week_data['content']['section_3_title']}</h2>
    <p>{week_data['content']['section_3_text']}</p>

    <h2>{week_data['content']['section_4_title']}</h2>
    <p>{week_data['content']['section_4_text']}</p>

    <p><strong>{week_data['content']['closing']}</strong></p>

    <div class="cta-buttons">
        <a href="{week_data['ctas'][0]['url']}" class="cta-button cta-primary">{week_data['ctas'][0]['text']}</a>
        <a href="{week_data['ctas'][1]['url']}" class="cta-button cta-secondary">{week_data['ctas'][1]['text']}</a>
    </div>

    <div class="footer">
        <h3>Research Citations</h3>
        <ul>'''

    for citation in week_data['metadata']['research_citations']:
        html += f'\n            <li>{citation}</li>'

    html += '''
        </ul>

        <h3>NGSS Alignment</h3>
        <ul>'''

    for practice in week_data['metadata']['ngss_practices']:
        html += f'\n            <li>{practice}</li>'

    html += f'''
        </ul>

        <p><strong>Authors:</strong> {week_data['author']}</p>
        <p><strong>Category:</strong> {week_data['category']}</p>
    </div>
</body>
</html>'''

    return html

# Create Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Blog Schedule"

# Set column widths
ws.column_dimensions['A'].width = 15  # Date
ws.column_dimensions['B'].width = 80  # Title
ws.column_dimensions['C'].width = 150 # HTML Content
ws.column_dimensions['D'].width = 15  # Status

# Add headers
headers = ['Date', 'Title', 'HTML Content', 'Status']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Starting date: Sunday, November 24, 2024
start_date = datetime(2024, 11, 24)

# Get GitHub username
username = get_github_username()

# Process all 52 weeks
week_json_dir = 'modelitk12-blog/content/week-json'
for week_num in range(1, 53):
    json_file = os.path.join(week_json_dir, f'week-{week_num:02d}-content.json')

    if not os.path.exists(json_file):
        print(f"Warning: {json_file} not found, skipping...")
        continue

    try:
        # Load the JSON file
        with open(json_file, 'r', encoding='utf-8') as f:
            week_data = json.load(f)

        # Get GitHub image URL
        image_url = get_github_image_url(username, week_num)

        # Convert to HTML
        week_html = json_to_simplified_html(week_data, image_url)

        # Calculate date for this week (each week is 7 days apart)
        week_date = start_date + timedelta(days=(week_num - 1) * 7)

        # Row number (header is row 1, data starts at row 2)
        row_num = week_num + 1

        # Add data to spreadsheet
        ws.cell(row=row_num, column=1).value = week_date.strftime('%Y-%m-%d')
        ws.cell(row=row_num, column=2).value = week_data['title']
        ws.cell(row=row_num, column=3).value = week_html
        ws.cell(row=row_num, column=3).alignment = Alignment(wrap_text=True, vertical='top')

        # Add status column (default to "Pending")
        ws.cell(row=row_num, column=4).value = "Pending"
        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center', vertical='top')

        # Set row height
        ws.row_dimensions[row_num].height = 30

        print(f"Processed Week {week_num}: {week_data['title'][:50]}...")

    except Exception as e:
        print(f"Error processing Week {week_num}: {str(e)}")

# Set header row height
ws.row_dimensions[1].height = 20

# Save the file
output_file = 'modelit-blog-schedule-final-with-html.xlsx'
wb.save(output_file)

print(f"\n{'='*60}")
print(f"SUCCESS: Complete blog schedule created!")
print(f"Output file: {output_file}")
print(f"Starting date: {start_date.strftime('%Y-%m-%d')} (Sunday)")
print(f"{'='*60}")
print(f"\nColumn Structure:")
print(f"  A: Date (weekly from Nov 24, 2024)")
print(f"  B: Title")
print(f"  C: HTML Content (complete with embedded images)")
print(f"  D: Status (default: 'Pending', change to 'Posted' after publishing)")
print(f"\nImages hosted at:")
print(f"  https://github.com/{username}/modelit-blog-images")
print(f"\nNext Steps:")
print(f"  1. Open this Excel file")
print(f"  2. Copy HTML from Column C to Blogger")
print(f"  3. Update Status to 'Posted' when complete")
print(f"  4. Or use Make.com to automate the entire process!")
