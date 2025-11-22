import json
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
import os
import html

def get_github_username():
    """Get GitHub username"""
    import subprocess
    result = subprocess.run(['gh', 'api', 'user', '--jq', '.login'], capture_output=True, text=True, check=True)
    return result.stdout.strip()

def get_github_image_url(username, week_num):
    """Generate GitHub raw URL for an image"""
    return f"https://raw.githubusercontent.com/{username}/modelit-blog-images/main/images/week-{week_num:02d}-hero.jpg"

def json_to_blogger_html(week_data, image_url):
    """Convert JSON to clean Blogger-ready HTML with beautiful image display"""

    # Escape HTML entities in text content to prevent quote issues
    title = html.escape(week_data['title'], quote=False)
    excerpt = html.escape(week_data['excerpt'], quote=False)
    opening = html.escape(week_data['content']['opening_hook'], quote=False)

    html_content = f'''<div style="max-width: 800px; margin: 0 auto; font-family: Arial, sans-serif; line-height: 1.6; color: #333;">

<!-- Hero Image with Beautiful Formatting -->
<div style="width: 100%; margin-bottom: 40px; text-align: center; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
    <img src="{image_url}" alt="{title}" style="width: 100%; height: auto; display: block; object-fit: cover; max-height: 500px;">
</div>

<!-- Title -->
<h1 style="color: #0066cc; font-size: 2.2em; margin-bottom: 20px; line-height: 1.3;">{title}</h1>

<!-- Excerpt -->
<div style="font-size: 1.2em; font-style: italic; color: #555; padding: 20px; background-color: #f5f5f5; border-left: 5px solid #0066cc; margin: 30px 0; border-radius: 4px;">
    <p style="margin: 0;">{excerpt}</p>
</div>

<!-- Opening Hook -->
<p style="font-size: 1.05em; line-height: 1.8; margin: 25px 0;">{opening}</p>

<!-- Section 1 -->
<h2 style="color: #0066cc; font-size: 1.6em; margin-top: 40px; margin-bottom: 15px;">{html.escape(week_data['content']['section_1_title'], quote=False)}</h2>
<p style="line-height: 1.8; margin: 20px 0;">{html.escape(week_data['content']['section_1_text'], quote=False)}</p>

<!-- Section 2 -->
<h2 style="color: #0066cc; font-size: 1.6em; margin-top: 40px; margin-bottom: 15px;">{html.escape(week_data['content']['section_2_title'], quote=False)}</h2>
<p style="line-height: 1.8; margin: 20px 0;">{html.escape(week_data['content']['section_2_text'], quote=False)}</p>

<!-- Section 3 -->
<h2 style="color: #0066cc; font-size: 1.6em; margin-top: 40px; margin-bottom: 15px;">{html.escape(week_data['content']['section_3_title'], quote=False)}</h2>
<p style="line-height: 1.8; margin: 20px 0;">{html.escape(week_data['content']['section_3_text'], quote=False)}</p>

<!-- Section 4 -->
<h2 style="color: #0066cc; font-size: 1.6em; margin-top: 40px; margin-bottom: 15px;">{html.escape(week_data['content']['section_4_title'], quote=False)}</h2>
<p style="line-height: 1.8; margin: 20px 0;">{html.escape(week_data['content']['section_4_text'], quote=False)}</p>

<!-- Closing -->
<div style="background-color: #f9f9f9; padding: 25px; margin: 40px 0; border-left: 5px solid #0066cc; border-radius: 4px;">
    <p style="line-height: 1.8; margin: 0; font-weight: 500;">{html.escape(week_data['content']['closing'], quote=False)}</p>
</div>

<!-- Call-to-Action Buttons -->
<div style="text-align: center; margin: 50px 0;">
    <a href="{week_data['ctas'][0]['url']}" style="display: inline-block; padding: 15px 35px; margin: 10px; background-color: #0066cc; color: white; text-decoration: none; font-weight: bold; border-radius: 6px; font-size: 1.1em; transition: background-color 0.3s;">{html.escape(week_data['ctas'][0]['text'], quote=False)}</a>
    <a href="{week_data['ctas'][1]['url']}" style="display: inline-block; padding: 15px 35px; margin: 10px; background-color: #28a745; color: white; text-decoration: none; font-weight: bold; border-radius: 6px; font-size: 1.1em; transition: background-color 0.3s;">{html.escape(week_data['ctas'][1]['text'], quote=False)}</a>
</div>

<!-- Footer -->
<div style="margin-top: 50px; padding: 30px; background-color: #f0f0f0; border-top: 3px solid #ccc; border-radius: 4px;">
    <h3 style="color: #333; font-size: 1.3em; margin-bottom: 15px;">Research Citations</h3>
    <ul style="line-height: 1.8; margin: 10px 0; padding-left: 25px;">'''

    for citation in week_data['metadata']['research_citations']:
        html_content += f'\n        <li>{html.escape(citation, quote=False)}</li>'

    html_content += '''
    </ul>

    <h3 style="color: #333; font-size: 1.3em; margin-top: 25px; margin-bottom: 15px;">NGSS Alignment</h3>
    <ul style="line-height: 1.8; margin: 10px 0; padding-left: 25px;">'''

    for practice in week_data['metadata']['ngss_practices']:
        html_content += f'\n        <li>{html.escape(practice, quote=False)}</li>'

    html_content += f'''
    </ul>

    <div style="margin-top: 25px; font-style: italic; color: #666;">
        <p style="margin: 5px 0;"><strong>Authors:</strong> {html.escape(week_data['author'], quote=False)}</p>
        <p style="margin: 5px 0;"><strong>Category:</strong> {html.escape(week_data['category'], quote=False)}</p>
    </div>
</div>

</div>'''

    return html_content

# Create Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Blog Schedule"

# Set column widths
ws.column_dimensions['A'].width = 15  # Date
ws.column_dimensions['B'].width = 80  # Title
ws.column_dimensions['C'].width = 150 # HTML Content
ws.column_dimensions['D'].width = 15  # Status

# Add headers
headers = ['Date', 'Title', 'HTML Content', 'Status']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Starting date: Sunday, November 24, 2024
start_date = datetime(2024, 11, 24)

# Get GitHub username
username = get_github_username()

# Process all 52 weeks
week_json_dir = 'modelitk12-blog/content/week-json'
for week_num in range(1, 53):
    json_file = os.path.join(week_json_dir, f'week-{week_num:02d}-content.json')

    if not os.path.exists(json_file):
        print(f"Warning: {json_file} not found, skipping...")
        continue

    try:
        # Load the JSON file
        with open(json_file, 'r', encoding='utf-8') as f:
            week_data = json.load(f)

        # Get GitHub image URL
        image_url = get_github_image_url(username, week_num)

        # Convert to clean HTML
        week_html = json_to_blogger_html(week_data, image_url)

        # Calculate date for this week
        week_date = start_date + timedelta(days=(week_num - 1) * 7)

        # Row number
        row_num = week_num + 1

        # Add data to spreadsheet
        ws.cell(row=row_num, column=1).value = week_date.strftime('%Y-%m-%d')
        ws.cell(row=row_num, column=2).value = week_data['title']
        ws.cell(row=row_num, column=3).value = week_html
        ws.cell(row=row_num, column=3).alignment = Alignment(wrap_text=True, vertical='top')

        # Add status column
        ws.cell(row=row_num, column=4).value = "Pending"
        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center', vertical='top')

        # Set row height
        ws.row_dimensions[row_num].height = 30

        print(f"SUCCESS: Week {week_num}: {week_data['title'][:60]}...")

    except Exception as e:
        print(f"Error processing Week {week_num}: {str(e)}")

# Set header row height
ws.row_dimensions[1].height = 20

# Save the file
output_file = 'modelit-blog-schedule-FIXED.xlsx'
wb.save(output_file)

print(f"\n{'='*70}")
print(f"SUCCESS: Fixed blog schedule created!")
print(f"Output file: {output_file}")
print(f"Starting date: {start_date.strftime('%Y-%m-%d')} (Sunday)")
print(f"{'='*70}")
print(f"\nFixes Applied:")
print(f"  ✓ Removed double quote escaping issues")
print(f"  ✓ Added beautiful image formatting with shadow and border-radius")
print(f"  ✓ Improved responsive image display")
print(f"  ✓ Enhanced overall visual styling")
print(f"  ✓ Clean HTML ready for Blogger")
print(f"\nImages hosted at:")
print(f"  https://github.com/{username}/modelit-blog-images")
